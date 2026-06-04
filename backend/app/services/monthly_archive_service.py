import os
import smtplib
from datetime import date, datetime, timedelta
from decimal import Decimal
from email.message import EmailMessage
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.car import Car
from app.models.debt import Debt
from app.models.inventory import InventoryItem
from app.models.invoice import Invoice
from app.models.service import Service
from app.models.tenant import Tenant
from app.models.user import Role, User


EXPORT_DIR = Path("/app/uploads/monthly_exports")
RETENTION_DAYS = 400
ENTERPRISE_PLAN = "enterprise"


def _plan_value(plan) -> str:
    return getattr(plan, "value", plan) or "basic"


def previous_month(today: date | None = None) -> tuple[int, int]:
    today = today or date.today()
    first_day = today.replace(day=1)
    last_prev_month = first_day - timedelta(days=1)
    return last_prev_month.year, last_prev_month.month


def _month_range(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)
    return start, end


def _value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _safe_filename(value: str) -> str:
    allowed = []
    for ch in value.strip():
        allowed.append(ch if ch.isalnum() else "-")
    cleaned = "".join(allowed).strip("-")
    return cleaned[:70] or "center"


def _append_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list]):
    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="E0F2FE")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        sheet.append([_value(item) for item in row])
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 3, 12), 42)


def _manager_email(db: Session, tenant_id: int) -> str | None:
    manager = db.query(User).filter(User.tenant_id == tenant_id, User.role == Role.manager).first()
    if not manager or manager.email.endswith("@carecar.app"):
        return None
    return manager.email


def export_tenant_monthly_archive(db: Session, tenant: Tenant, year: int, month: int) -> dict:
    start, end = _month_range(year, month)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    cars = db.query(Car).filter(Car.tenant_id == tenant.id).order_by(Car.created_at.desc()).all()
    services = db.query(Service).filter(
        Service.tenant_id == tenant.id,
        Service.service_date >= start,
        Service.service_date < end,
    ).order_by(Service.service_date.desc(), Service.id.desc()).all()
    invoices = db.query(Invoice).filter(
        Invoice.tenant_id == tenant.id,
        Invoice.invoice_date >= start,
        Invoice.invoice_date < end,
    ).order_by(Invoice.invoice_date.desc(), Invoice.id.desc()).all()
    debts = db.query(Debt).filter(Debt.tenant_id == tenant.id, Debt.amount > 0).order_by(Debt.created_at.desc()).all()
    inventory = db.query(InventoryItem).filter(InventoryItem.tenant_id == tenant.id).order_by(InventoryItem.oil_type).all()

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Care Car Monthly Archive"])
    summary.append(["Center", tenant.name])
    summary.append(["Period", f"{year}-{month:02d}"])
    summary.append(["Generated at", datetime.now().isoformat(timespec="seconds")])
    summary.append([])
    summary.append(["Cars total", len(cars)])
    summary.append(["Services this month", len(services)])
    summary.append(["Invoices this month", len(invoices)])
    summary.append(["Open debts", len(debts)])
    summary.append(["Inventory items", len(inventory)])
    summary["A1"].font = Font(bold=True, size=16)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 36

    _append_sheet(workbook, "Cars", [
        "ID", "Plate", "Owner", "Phone", "Type", "Color", "Notes", "Created",
    ], [[c.id, c.plate_number, c.owner_name, c.phone, c.car_type, c.car_color, c.notes, c.created_at] for c in cars])

    car_by_id = {car.id: car for car in cars}
    _append_sheet(workbook, "Services", [
        "ID", "Date", "Plate", "Owner", "Service", "Mileage", "Notes", "Created",
    ], [[
        s.id, s.service_date, car_by_id.get(s.car_id).plate_number if car_by_id.get(s.car_id) else "",
        car_by_id.get(s.car_id).owner_name if car_by_id.get(s.car_id) else "",
        s.oil_type, s.mileage, s.notes, s.created_at,
    ] for s in services])

    service_by_id = {service.id: service for service in services}
    _append_sheet(workbook, "Invoices", [
        "ID", "Date", "Plate", "Service", "Amount", "Discount", "Net", "Status", "Created",
    ], [[
        inv.id, inv.invoice_date,
        car_by_id.get(service_by_id.get(inv.service_id).car_id).plate_number if service_by_id.get(inv.service_id) and car_by_id.get(service_by_id.get(inv.service_id).car_id) else "",
        service_by_id.get(inv.service_id).oil_type if service_by_id.get(inv.service_id) else "",
        inv.amount, inv.discount, float(inv.amount or 0) - float(inv.discount or 0), inv.status.value if hasattr(inv.status, "value") else inv.status,
        inv.created_at,
    ] for inv in invoices])

    _append_sheet(workbook, "Open Debts", [
        "ID", "Invoice ID", "Plate", "Amount", "Due Date", "Notes", "Created",
    ], [[
        debt.id, debt.invoice_id, car_by_id.get(debt.car_id).plate_number if car_by_id.get(debt.car_id) else "",
        debt.amount, debt.due_date, debt.notes, debt.created_at,
    ] for debt in debts])

    _append_sheet(workbook, "Inventory", [
        "ID", "Product", "Category", "Supplier", "Quantity", "Min", "Cost", "Sale", "Sold",
    ], [[
        item.id, item.oil_type, item.category, item.supplier_name, item.quantity,
        item.min_threshold, item.unit_cost, item.sale_price, item.total_sold,
    ] for item in inventory])

    filename = f"{tenant.id}-{_safe_filename(tenant.name)}-{year}-{month:02d}.xlsx"
    path = EXPORT_DIR / filename
    workbook.save(path)
    return {
        "tenant_id": tenant.id,
        "tenant_name": tenant.name,
        "filename": filename,
        "path": str(path),
        "url": f"{settings.PUBLIC_BASE_URL.rstrip('/')}/uploads/monthly_exports/{filename}",
        "counts": {
            "cars": len(cars),
            "services": len(services),
            "invoices": len(invoices),
            "debts": len(debts),
            "inventory": len(inventory),
        },
    }


def send_archive_email(to_email: str, tenant: Tenant, archive: dict, year: int, month: int) -> str:
    if not settings.SMTP_HOST or not settings.SMTP_USER or not settings.SMTP_PASSWORD:
        return "not_configured"

    message = EmailMessage()
    message["Subject"] = f"نسخة بيانات {tenant.name} الشهرية - {year}-{month:02d}"
    message["From"] = f"{settings.SMTP_FROM_NAME} <{settings.SMTP_FROM_EMAIL or settings.SMTP_USER}>"
    message["To"] = to_email
    message.set_content("\n".join([
        f"مرحباً {tenant.name}",
        f"مرفق نسخة Excel من بيانات مركزك لشهر {year}-{month:02d}.",
        "هذه النسخة للمتابعة والأرشفة فقط.",
        archive["url"],
    ]))
    with open(archive["path"], "rb") as handle:
        message.add_attachment(
            handle.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=archive["filename"],
        )

    try:
        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            smtp.send_message(message)
        return "sent"
    except Exception as exc:
        return f"failed: {exc}"


def send_archive_whatsapp(phone: str | None, tenant: Tenant, archive: dict, year: int, month: int) -> str:
    if not settings.PLATFORM_WASNDER_API_KEY or not settings.PLATFORM_WHATSAPP_NUMBER:
        return "not_configured"
    if not phone:
        return "missing_phone"

    recipient = phone.strip()
    if recipient.startswith("0"):
        recipient = "+964" + recipient[1:]

    message = "\n".join([
        f"مرحباً {tenant.name}",
        f"هذه نسخة Excel الشهرية لبيانات مركزك عن {year}-{month:02d}.",
        "رابط التحميل:",
        archive["url"],
        "احتفظ بها للأرشفة والمتابعة.",
    ])
    try:
        response = httpx.post(
            settings.WASNDER_API_URL,
            json={"to": recipient, "text": message},
            headers={"Authorization": f"Bearer {settings.PLATFORM_WASNDER_API_KEY}"},
            timeout=15,
        )
        return "sent" if response.is_success else f"failed: {response.text[:300]}"
    except httpx.HTTPError as exc:
        return f"failed: {exc}"


def cleanup_old_monthly_archives(retention_days: int = RETENTION_DAYS) -> int:
    if not EXPORT_DIR.exists():
        return 0
    cutoff = datetime.now() - timedelta(days=retention_days)
    removed = 0
    for path in EXPORT_DIR.glob("*.xlsx"):
        modified = datetime.fromtimestamp(path.stat().st_mtime)
        if modified < cutoff:
            path.unlink(missing_ok=True)
            removed += 1
    return removed


def run_monthly_archives(db: Session, year: int | None = None, month: int | None = None) -> dict:
    if not year or not month:
        year, month = previous_month()

    tenants = db.query(Tenant).filter(Tenant.is_active == True).order_by(Tenant.name).all()
    results = []
    for tenant in tenants:
        if _plan_value(tenant.plan) != ENTERPRISE_PLAN:
            continue
        archive = export_tenant_monthly_archive(db, tenant, year, month)
        email = _manager_email(db, tenant.id)
        phone = tenant.whatsapp_number or tenant.contact_phone
        results.append({
            **archive,
            "email_status": send_archive_email(email, tenant, archive, year, month) if email else "missing_email",
            "whatsapp_status": send_archive_whatsapp(phone, tenant, archive, year, month) if phone else "missing_phone",
        })

    removed = cleanup_old_monthly_archives()
    return {
        "year": year,
        "month": month,
        "tenant_count": len(results),
        "removed_old_files": removed,
        "results": results,
    }
